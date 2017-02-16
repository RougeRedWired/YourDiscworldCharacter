""" Frequency in a string is determined by the number of occurences/length of the string
So first count the number of occurences for a letter and then divide it by len(string)Return the frequency
"""
import xlsxwriter
import openpyxl

#data handling function
def load_index_to_xls(filename,index):
   
#create the point handling function
   def score(string):
      score = 5
      if string.find(letter)!=-1:
         if string[letter][frequency] == index[string][letter][frequency]:
            score +=5
         elif (string[letter][frequency] - index[string][letter][frequency] <= 1):
         score +=3

         elif (string[letter][frequency] - index[string][letter][frequency] > 1):
         score +=1

      return score
      
         
         

#Create a file to put the data in
def create_index_file():
   workbook = xlsxwriter.Workbook('index.xlsx')
   worksheet = workbook.add_worksheet()
   worksheet.write('A1', 'Word')
   workbook.close() 


#Get frequency of a letter

def frequency_return(string,letter):
    count=0
    ranging_string=len(string)
    for letters in string:
        
       if letters==letter:
           count+=1
    return count

index={}

from openpyxl import load_workbook
index = load_workbook('index.xlsx')




#Scan all letters: if a letter has not been searched then count

def store_words_and_frequency(string):
    range_string=string
    length_string=len(string)
    datastore=[]
    target=0
    frequency=0
    
    if string not in index:
        while len(range_string)!=0:
            datastore.append(range_string[target])
            frequency = (int(frequency_return(range_string,range_string[target]))/length_string)
            datastore.append(frequency)
            range_string = range_string.replace(range_string[target],'')
            index.update({string: datastore})
        return index
    else:
        return index
    
    
store_words_and_frequency("string")
store_words_and_frequency("string")
store_words_and_frequency("boat")
store_words_and_frequency("house")

print(index)







    
    
   
        
